from __future__ import annotations

import csv
import re
import unicodedata
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from pypdf import PdfReader
from sqlalchemy import text

from app.core.exceptions import ApiServiceError
from app.db.session import get_connection, get_transaction
from app.services import ledger_service, purchase_pdf_service
from app.utils.company_access import build_in_clause, can_access_company, user_company_ids


DEFAULT_IVA_RATE = 16.0
PROVIDER_NAME_SQL = "COALESCE(NULLIF(cl.CommercialName, ''), cl.LegalName)"
ALLOWED_INVOICE_CONTENT_TYPES = {
    "application/pdf": ".pdf",
    "application/xml": ".xml",
    "text/xml": ".xml",
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
}
SHEET_FIELD_ALIASES = {
    "codigo": ["codigo", "clave", "sku", "item", "articulo", "producto", "material", "codigoproveedor"],
    "descripcion": [
        "descripcion",
        "concepto",
        "producto descripcion",
        "articulo descripcion",
        "nombre",
        "material descripcion",
        "descripcion material",
    ],
    "cantidad": ["cantidad", "cant", "qty", "unidades", "kgs", "kg", "cantidad facturada"],
    "precio": [
        "precio",
        "precio unitario",
        "costo",
        "costo unitario",
        "unit price",
        "importe unitario",
        "valor unitario",
    ],
    "iva": ["iva", "iva%", "impuesto", "tasa iva", "porcentaje iva"],
}
NOISE_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in [
        r"subtotal",
        r"^total$",
        r"importe total",
        r"folio fiscal",
        r"fecha.*certif",
        r"metodo de pago",
        r"forma de pago",
        r"cadena original",
        r"sello digital",
        r"numero de serie",
        r"uso del cfdi",
        r"regimen fiscal",
        r"tipo comprobante",
        r"lugar de expedicion",
        r"^r\.f\.c\.",
        r"^rfc:",
        r"domicilio fiscal",
        r"^serie:",
        r"^folio:",
        r"^fecha y hora",
        r"observaciones",
        r"favor de pagar",
        r"transferencia bancaria",
        r"^impuesto:",
        r"^descripci",
        r"^cantidadunidad",
        r"^clav[e]",
        r"^sat$",
        r"^unidad$",
        r"\|\|\|1\.1\|",
    ]
]
BASE_DIR = Path(__file__).resolve().parents[2]
UPLOADS_DIR = BASE_DIR / "uploads" / "compras-facturas"


def _username(current_user: dict[str, Any]) -> str:
    return str(
        current_user.get("Nombre")
        or current_user.get("Name")
        or current_user.get("Username")
        or current_user.get("username")
        or current_user.get("Email")
        or "Sistema"
    )


def _normalize_text(value: Any) -> str:
    normalized = unicodedata.normalize("NFD", str(value or ""))
    without_accents = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", without_accents.lower()).strip()


def _parse_numeric_value(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)

    raw = re.sub(r"[^\d,.\-]", "", str(value or "").strip())
    if not raw:
        return 0.0

    last_comma = raw.rfind(",")
    last_dot = raw.rfind(".")
    if last_comma > -1 and last_dot > -1:
        raw = raw.replace(".", "").replace(",", ".") if last_comma > last_dot else raw.replace(",", "")
    elif last_comma > -1:
        raw = raw.replace(",", ".") if raw.count(",") == 1 else raw.replace(",", "")

    try:
        return float(raw)
    except ValueError:
        return 0.0


def _tokenize(value: Any) -> list[str]:
    return [token for token in _normalize_text(value).split(" ") if len(token) > 2]


def _detect_sheet_fields(header_row: list[Any]) -> dict[str, int]:
    normalized_headers = [_normalize_text(header) for header in header_row]
    mapping: dict[str, int] = {}
    for field, aliases in SHEET_FIELD_ALIASES.items():
        for index, header in enumerate(normalized_headers):
            if header in aliases:
                mapping[field] = index
                break
    return mapping


def _sheet_matrix_from_bytes(file_bytes: bytes, extension: str) -> list[list[Any]]:
    if extension == ".csv":
        decoded = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                decoded = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            decoded = file_bytes.decode("utf-8", errors="ignore")
        return [row for row in csv.reader(StringIO(decoded)) if any(str(cell).strip() for cell in row)]

    if extension == ".xlsx":
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True) if any(cell not in (None, "") for cell in row)]

    if extension == ".xls":
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        rows: list[list[Any]] = []
        for index in range(sheet.nrows):
            row = sheet.row_values(index)
            if any(str(cell).strip() for cell in row):
                rows.append(row)
        return rows

    raise ApiServiceError(
        status_code=400,
        content={"success": False, "message": "Formato de hoja no soportado"},
    )


def _extract_rows_from_supplier_sheet(file_bytes: bytes, extension: str) -> list[dict[str, Any]]:
    matrix = _sheet_matrix_from_bytes(file_bytes, extension)
    if not matrix:
        return []

    best_header_index = 0
    best_score = -1
    for index, row in enumerate(matrix[:10]):
        score = len(_detect_sheet_fields([str(cell or "") for cell in row]))
        if score > best_score:
            best_score = score
            best_header_index = index

    header_row = [str(cell or "") for cell in matrix[best_header_index]]
    mapping = _detect_sheet_fields(header_row)

    rows: list[dict[str, Any]] = []
    for offset, row in enumerate(matrix[best_header_index + 1 :], start=best_header_index + 2):
        codigo = str(row[mapping["codigo"]]).strip() if "codigo" in mapping and mapping["codigo"] < len(row) else ""
        descripcion = (
            str(row[mapping["descripcion"]]).strip()
            if "descripcion" in mapping and mapping["descripcion"] < len(row)
            else ""
        )
        cantidad = (
            _parse_numeric_value(row[mapping["cantidad"]])
            if "cantidad" in mapping and mapping["cantidad"] < len(row)
            else 0
        )
        precio = (
            _parse_numeric_value(row[mapping["precio"]])
            if "precio" in mapping and mapping["precio"] < len(row)
            else 0
        )
        iva = (
            _parse_numeric_value(row[mapping["iva"]])
            if "iva" in mapping and mapping["iva"] < len(row)
            else DEFAULT_IVA_RATE
        )
        item = {
            "rowNumber": offset,
            "codigo": codigo,
            "descripcion": descripcion,
            "cantidad": cantidad or 0,
            "precio": precio or 0,
            "iva": iva or DEFAULT_IVA_RATE,
        }
        if item["codigo"] or item["descripcion"] or item["cantidad"] > 0 or item["precio"] > 0:
            rows.append(item)
    return rows


def _is_numeric_token(token: Any) -> bool:
    return bool(re.match(r"^[-$]?[\d.,]+%?$", str(token or "").strip()))


def _is_noise_line(line: str) -> bool:
    normalized = _normalize_text(line)
    if len(normalized) < 2:
        return True
    if re.match(r"^[0-9a-f]{8}-[0-9a-f]{4}", line, re.IGNORECASE):
        return True
    if re.match(r"^\d{17,}$", normalized.replace(" ", "")):
        return True
    return any(pattern.search(line) for pattern in NOISE_PATTERNS)


def _parse_smurfit_blocks(lines: list[str]) -> list[dict[str, Any]]:
    block_start = re.compile(r"^(\d+\.\d+)\d{7,}(?:TO|KG|LT|PZ|UN)?$", re.IGNORECASE)
    price_line = re.compile(r"^\$([\d,]+\.\d{2})\$([\d,]+\.\d{2})$")
    results: list[dict[str, Any]] = []
    index = 0
    while index < len(lines):
        start_match = block_start.match(lines[index])
        if not start_match:
            index += 1
            continue

        qty = float(start_match.group(1))
        description_parts: list[str] = []
        precio = 0.0
        cursor = index + 1
        while cursor < len(lines) and cursor < index + 10:
            stripped = lines[cursor].replace(" ", "")
            price_match = price_line.match(stripped)
            if price_match:
                precio = _parse_numeric_value(price_match.group(1))
                cursor += 1
                break
            if block_start.match(lines[cursor]):
                break
            if not _is_noise_line(lines[cursor]) and len(lines[cursor]) > 2:
                description_parts.append(lines[cursor].strip())
            cursor += 1

        descripcion = " ".join(description_parts).strip()
        if descripcion and qty > 0:
            results.append(
                {
                    "rowNumber": index + 1,
                    "codigo": "",
                    "descripcion": descripcion,
                    "cantidad": qty,
                    "precio": precio,
                    "iva": DEFAULT_IVA_RATE,
                }
            )
        index = cursor
    return results


def _parse_univalix_blocks(lines: list[str]) -> list[dict[str, Any]]:
    row_start = re.compile(r"^(\d+(?:\.\d+)?)0{3,}[\w-]+")
    results: list[dict[str, Any]] = []
    index = 0
    while index < len(lines):
        if not row_start.match(lines[index]):
            index += 1
            continue

        qty_match = re.match(r"^(\d+(?:\.\d+)?)", lines[index])
        qty = float(qty_match.group(1)) if qty_match else 0.0
        code_match = re.search(r"0{3,}([\w-]+)$", lines[index])
        supplier_code = code_match.group(1) if code_match else ""
        description_line = lines[index + 1].strip() if index + 1 < len(lines) else ""
        if description_line:
            parts = description_line.split()
            desc_end = len(parts)
            for cursor in range(len(parts) - 1, 0, -1):
                if not _is_numeric_token(parts[cursor]):
                    desc_end = cursor + 1
                    break
            descripcion = " ".join(parts[:desc_end]).strip() or description_line
            numeric_values = [_parse_numeric_value(part) for part in parts[desc_end:] if _parse_numeric_value(part) > 0]
            precio = numeric_values[-2] if len(numeric_values) >= 2 else (numeric_values[0] if numeric_values else 0)
            results.append(
                {
                    "rowNumber": index + 1,
                    "codigo": supplier_code,
                    "descripcion": descripcion,
                    "cantidad": qty if qty > 0 else 1,
                    "precio": precio,
                    "iva": DEFAULT_IVA_RATE,
                }
            )
        index += 3
    return results


def _parse_adhempa_blocks(lines: list[str]) -> list[dict[str, Any]]:
    dense = re.compile(
        r"^([\d,]+(?:\.\d+)?)([A-Z]{1,5}\s*\d{1,5}(?:\s*[A-Z]{0,5}\s*\d{0,5})*)(.*?)([\d,]+\.\d{2})[A-Z]"
    )
    results: list[dict[str, Any]] = []
    for index, line in enumerate(lines, start=1):
        match = dense.match(line)
        if not match:
            continue
        qty = _parse_numeric_value(match.group(1))
        codigo = re.sub(r"\s+", " ", match.group(2)).strip()
        total_price = _parse_numeric_value(match.group(4))
        if qty > 0 and total_price > 0:
            results.append(
                {
                    "rowNumber": index,
                    "codigo": codigo,
                    "descripcion": codigo,
                    "cantidad": qty,
                    "precio": round(total_price / qty, 2),
                    "iva": DEFAULT_IVA_RATE,
                }
            )
    return results


def _parse_generic_pdf_lines(lines: list[str]) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for index, line in enumerate(lines, start=1):
        if _is_noise_line(line):
            continue
        tokens = line.split()
        numeric_tokens = [_parse_numeric_value(token) for token in tokens if _is_numeric_token(token)]
        numeric_values = [value for value in numeric_tokens if 0 < value < 1e9]
        text_tokens = [token for token in tokens if not _is_numeric_token(token) and len(token) > 2 and re.search(r"[a-zA-Z]", token)]
        if len(numeric_values) < 1 or len(text_tokens) < 1:
            continue
        if len(numeric_values) == 1 and numeric_values[0] > 10000:
            continue
        qty = numeric_values[0]
        precio = numeric_values[-1] if len(numeric_values) >= 2 else 0
        descripcion = " ".join(text_tokens).strip()
        code_candidate = tokens[0] if tokens else ""
        codigo = code_candidate if re.match(r"^[A-Z0-9][\w-]{1,20}$", code_candidate) else ""
        if descripcion:
            results.append(
                {
                    "rowNumber": index,
                    "codigo": codigo,
                    "descripcion": descripcion,
                    "cantidad": qty if qty > 0 else 1,
                    "precio": precio,
                    "iva": DEFAULT_IVA_RATE,
                }
            )
    return results


def _extract_rows_from_supplier_pdf(file_bytes: bytes) -> list[dict[str, Any]]:
    reader = PdfReader(BytesIO(file_bytes))
    text_content = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = [line.strip() for line in text_content.splitlines() if line.strip()]

    if len([line for line in lines if re.match(r"^\d+\.\d+\d{7,}(?:TO|KG|LT|PZ|UN)?$", line, re.IGNORECASE)]) >= 2:
        rows = _parse_smurfit_blocks(lines)
        if rows:
            return rows

    if len([line for line in lines if re.match(r"^\d+(?:\.\d+)?0{3,}[\w-]", line)]) >= 1:
        rows = _parse_univalix_blocks(lines)
        if rows:
            return rows

    if len([line for line in lines if re.match(r"^[\d,]+[A-Z]", line) and len(line) > 20]) >= 1:
        rows = _parse_adhempa_blocks(lines)
        if rows:
            return rows

    return _parse_generic_pdf_lines(lines)


def _extract_rows_from_supplier_document(file_bytes: bytes, extension: str) -> list[dict[str, Any]]:
    return _extract_rows_from_supplier_pdf(file_bytes) if extension == ".pdf" else _extract_rows_from_supplier_sheet(file_bytes, extension)


def _build_materia_prima_index(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    indexed = []
    for row in rows:
        indexed.append(
            {
                **dict(row),
                "codigoNorm": _normalize_text(row.get("Codigo")),
                "nombreNorm": _normalize_text(row.get("Nombre")),
                "descripcionNorm": _normalize_text(row.get("Descripcion")),
                "tokens": list(
                    {
                        *(_tokenize(row.get("Codigo"))),
                        *(_tokenize(row.get("Nombre"))),
                        *(_tokenize(row.get("Descripcion"))),
                    }
                ),
            }
        )
    return indexed


def _find_best_materia_prima_match(
    source_row: dict[str, Any],
    materia_prima_index: list[dict[str, Any]],
) -> dict[str, Any] | None:
    codigo_norm = _normalize_text(source_row.get("codigo"))
    descripcion_norm = _normalize_text(source_row.get("descripcion"))
    source_tokens = list({*(_tokenize(source_row.get("codigo"))), *(_tokenize(source_row.get("descripcion")))})

    best: dict[str, Any] | None = None
    for materia_prima in materia_prima_index:
        score = 0
        if codigo_norm and materia_prima["codigoNorm"] == codigo_norm:
            score += 120
        if codigo_norm and materia_prima["codigoNorm"] and (
            materia_prima["codigoNorm"] in codigo_norm or codigo_norm in materia_prima["codigoNorm"]
        ):
            score += 45
        if descripcion_norm and materia_prima["nombreNorm"] == descripcion_norm:
            score += 100
        if descripcion_norm and materia_prima["descripcionNorm"] == descripcion_norm:
            score += 85
        if descripcion_norm and materia_prima["nombreNorm"] and (
            materia_prima["nombreNorm"] in descripcion_norm or descripcion_norm in materia_prima["nombreNorm"]
        ):
            score += 55
        if descripcion_norm and materia_prima["descripcionNorm"] and (
            materia_prima["descripcionNorm"] in descripcion_norm or descripcion_norm in materia_prima["descripcionNorm"]
        ):
            score += 40

        token_matches = len([token for token in source_tokens if token in materia_prima["tokens"]])
        score += token_matches * 12

        if not best or score > best["score"]:
            best = {"materiaPrima": materia_prima, "score": score}

    if not best or best["score"] < 36:
        return None

    return {
        **best["materiaPrima"],
        "confidence": min(100, round((best["score"] / 120) * 100)),
    }


def _next_numero_oc(connection: Any, company_id: int) -> str:
    row = connection.execute(
        text(
            """
            SELECT COUNT(*) AS Total
            FROM ERP_COMPRA_ORDEN
            WHERE Company_Id = :company_id
            """
        ),
        {"company_id": company_id},
    ).mappings().first()
    sequence = int((row or {}).get("Total") or 0) + 1
    return f"OC-{company_id:03d}-{sequence:05d}"


def _calculate_purchase_totals(items: list[dict[str, Any]]) -> tuple[float, float, float]:
    subtotal = 0.0
    total_iva = 0.0
    for item in items:
        line_subtotal = float(item.get("Cantidad") or 0) * float(item.get("PrecioCompra") or 0)
        iva = line_subtotal * (float(item.get("IVA", DEFAULT_IVA_RATE) or DEFAULT_IVA_RATE) / 100)
        subtotal += line_subtotal
        total_iva += iva
    return subtotal, total_iva, subtotal + total_iva


def _ensure_purchase_company_access(current_user: dict[str, Any], company_id: int | None) -> None:
    if not can_access_company(current_user, company_id):
        raise ApiServiceError(
            status_code=403,
            content={"success": False, "message": "No tiene permisos para operar compras de esta empresa"},
        )


def _load_purchase_order(connection: Any, order_id: int) -> dict[str, Any] | None:
    row = connection.execute(
        text(
            f"""
            SELECT
                oc.*,
                c.NameCompany AS Empresa,
                {PROVIDER_NAME_SQL} AS Proveedor,
                cl.RFC AS ProveedorRFC,
                contact.Email AS ProveedorEmail,
                req.Req_Id AS Requisicion_Id,
                req.NumeroReq AS RequisicionNumero
            FROM ERP_COMPRA_ORDEN oc
            LEFT JOIN ERP_COMPANY c ON oc.Company_Id = c.Company_Id
            LEFT JOIN ERP_CLIENT cl ON oc.Proveedor_Id = cl.Client_Id
            OUTER APPLY (
                SELECT TOP 1 cc.Email
                FROM ERP_CLIENTCONTACTS cc
                WHERE cc.Client_Id = cl.Client_Id
                ORDER BY cc.IsPrimary DESC, cc.Contact_Id ASC
            ) contact
            OUTER APPLY (
                SELECT TOP 1 r.Req_Id, r.NumeroReq
                FROM ERP_REQUISICION_COMPRA r
                WHERE r.OC_Id = oc.OC_Id
                ORDER BY r.Req_Id DESC
            ) req
            WHERE oc.OC_Id = :oc_id
            """
        ),
        {"oc_id": order_id},
    ).mappings().first()
    return dict(row) if row else None


def _load_purchase_details(connection: Any, order_id: int) -> list[dict[str, Any]]:
    rows = connection.execute(
        text(
            """
            SELECT
                d.*,
                COALESCE(p.SKU, '') AS ProductoSKU,
                COALESCE(p.Nombre, '') AS ProductoNombre,
                COALESCE(mp.Nombre, '') AS MateriaPrimaNombre
            FROM ERP_COMPRA_ORDEN_DETALLE d
            LEFT JOIN ERP_PRODUCTOS p ON d.Producto_Id = p.Producto_Id
            LEFT JOIN ERP_MATERIA_PRIMA mp ON d.MateriaPrima_Id = mp.MateriaPrima_Id
            WHERE d.OC_Id = :oc_id
            """
        ),
        {"oc_id": order_id},
    ).mappings().all()
    return [dict(row) for row in rows]


def _load_purchase_authorizations(connection: Any, order_id: int) -> list[dict[str, Any]]:
    rows = connection.execute(
        text(
            """
            SELECT
                a.*,
                LTRIM(ISNULL(u.Name, '') + ' ' + ISNULL(u.Lastname, '')) AS AutorizadoPor
            FROM ERP_COMPRA_AUTORIZACION a
            LEFT JOIN ERP_USERS u ON a.User_Id = u.User_Id
            WHERE a.OC_Id = :oc_id
            ORDER BY a.Nivel
            """
        ),
        {"oc_id": order_id},
    ).mappings().all()
    return [dict(row) for row in rows]


def analyze_supplier_sheet_direct(file_bytes: bytes, filename: str) -> dict[str, Any]:
    extension = Path(filename or "").suffix.lower()
    if extension not in {".xlsx", ".xls", ".csv", ".pdf"}:
        raise ApiServiceError(
            status_code=400,
            content={
                "success": False,
                "message": "Solo se permiten archivos Excel, CSV o PDF (.xlsx, .xls, .csv, .pdf)",
            },
        )

    source_rows = _extract_rows_from_supplier_document(file_bytes, extension)
    if not source_rows:
        raise ApiServiceError(
            status_code=400,
            content={
                "success": False,
                "message": "No se detectaron renglones utilizables en el archivo del proveedor",
            },
        )

    with get_connection() as connection:
        materia_prima_rows = connection.execute(
            text(
                """
                SELECT MateriaPrima_Id, Codigo, Nombre, Descripcion, CostoUnitario, Moneda
                FROM ERP_MATERIA_PRIMA
                WHERE ISNULL(Activo, 1) = 1
                ORDER BY Nombre
                """
            )
        ).mappings().all()

    materia_prima_index = _build_materia_prima_index([dict(row) for row in materia_prima_rows])
    items: list[dict[str, Any]] = []
    for row in source_rows:
        match = _find_best_materia_prima_match(row, materia_prima_index)
        precio_compra = float(row.get("precio") or 0) or float((match or {}).get("CostoUnitario") or 0)
        if match:
            items.append(
                {
                    "Tipo": "mp",
                    "MateriaPrima_Id": match.get("MateriaPrima_Id"),
                    "Producto_Id": "",
                    "Descripcion": match.get("Nombre") or row.get("descripcion") or row.get("codigo"),
                    "Cantidad": float(row.get("cantidad") or 0) or 1,
                    "PrecioCompra": precio_compra,
                    "IVA": float(row.get("iva") or 0) or DEFAULT_IVA_RATE,
                    "ReferenciaProveedor": row.get("codigo") or "",
                    "DescripcionProveedor": row.get("descripcion") or "",
                    "MatchNombre": match.get("Nombre"),
                    "MatchCodigo": match.get("Codigo"),
                    "MatchConfidence": match.get("confidence"),
                    "RowNumber": row.get("rowNumber"),
                }
            )
        else:
            items.append(
                {
                    "Tipo": "otro",
                    "MateriaPrima_Id": "",
                    "Producto_Id": "",
                    "Descripcion": row.get("descripcion") or row.get("codigo") or f"Renglon {row.get('rowNumber')}",
                    "Cantidad": float(row.get("cantidad") or 0) or 1,
                    "PrecioCompra": precio_compra,
                    "IVA": float(row.get("iva") or 0) or DEFAULT_IVA_RATE,
                    "ReferenciaProveedor": row.get("codigo") or "",
                    "DescripcionProveedor": row.get("descripcion") or "",
                    "MatchConfidence": 0,
                    "RowNumber": row.get("rowNumber"),
                }
            )

    materias_ligadas = len([item for item in items if item["Tipo"] == "mp"])
    return {
        "success": True,
        "data": {
            "items": items,
            "resumen": {
                "archivo": filename,
                "lineasDetectadas": len(items),
                "materiasLigadas": materias_ligadas,
                "lineasPendientes": len(items) - materias_ligadas,
            },
        },
    }


def list_providers(filters: dict[str, Any], current_user: dict[str, Any]) -> dict[str, Any]:
    request_company_id = int(filters["Company_Id"]) if filters.get("Company_Id") else None
    if request_company_id:
        _ensure_purchase_company_access(current_user, request_company_id)

    joins = ""
    where_clauses = ["c.ClientType IN ('PROVEEDOR','AMBOS')"]
    params: dict[str, Any] = {}

    if request_company_id:
        joins = " INNER JOIN ERP_CLIENTCOMPANIES cc ON c.Client_Id = cc.Client_Id"
        where_clauses.append("cc.Company_Id = :company_id")
        params["company_id"] = request_company_id
    elif not current_user.get("is_admin"):
        companies = user_company_ids(current_user)
        if not companies:
            return {"success": True, "data": []}
        clause, clause_params = build_in_clause("company", companies)
        joins = " INNER JOIN ERP_CLIENTCOMPANIES cc ON c.Client_Id = cc.Client_Id"
        where_clauses.append(f"cc.Company_Id IN ({clause})")
        params.update(clause_params)

    with get_connection() as connection:
        rows = connection.execute(
            text(
                f"""
                SELECT DISTINCT
                    c.Client_Id,
                    c.LegalName,
                    c.CommercialName,
                    COALESCE(NULLIF(c.CommercialName, ''), c.LegalName) AS ProviderName,
                    c.RFC,
                    contact.Email,
                    c.ClientType
                FROM ERP_CLIENT c
                {joins}
                OUTER APPLY (
                    SELECT TOP 1 ct.Email
                    FROM ERP_CLIENTCONTACTS ct
                    WHERE ct.Client_Id = c.Client_Id
                    ORDER BY ct.IsPrimary DESC, ct.Contact_Id ASC
                ) contact
                WHERE {' AND '.join(where_clauses)}
                ORDER BY COALESCE(NULLIF(c.CommercialName, ''), c.LegalName)
                """
            ),
            params,
        ).mappings().all()
    return {"success": True, "data": [dict(row) for row in rows]}


def list_purchase_orders(filters: dict[str, Any], current_user: dict[str, Any]) -> dict[str, Any]:
    company_id = int(filters["Company_Id"]) if filters.get("Company_Id") else None
    if company_id:
        _ensure_purchase_company_access(current_user, company_id)

    query = f"""
        SELECT
            oc.OC_Id,
            oc.NumeroOC,
            oc.Company_Id,
            c.NameCompany AS Empresa,
            oc.Proveedor_Id,
            {PROVIDER_NAME_SQL} AS Proveedor,
            oc.FechaOC,
            oc.FechaRequerida,
            oc.Moneda,
            oc.Subtotal,
            oc.IVA,
            oc.Total,
            oc.Estatus,
            oc.RequiereDobleAutorizacion,
            oc.FacturaReferencia,
            oc.FacturaArchivoUrl,
            oc.Observaciones,
            oc.CreatedBy,
            oc.CreatedAt,
            oc.UpdatedAt,
            (
                SELECT COUNT(*)
                FROM ERP_COMPRA_AUTORIZACION a
                WHERE a.OC_Id = oc.OC_Id AND a.Aprobado = 1
            ) AS AutorizacionesOtorgadas
        FROM ERP_COMPRA_ORDEN oc
        LEFT JOIN ERP_COMPANY c ON oc.Company_Id = c.Company_Id
        LEFT JOIN ERP_CLIENT cl ON oc.Proveedor_Id = cl.Client_Id
        WHERE 1 = 1
    """
    params: dict[str, Any] = {}
    if company_id:
        query += " AND oc.Company_Id = :company_id"
        params["company_id"] = company_id
    elif not current_user.get("is_admin"):
        companies = user_company_ids(current_user)
        if not companies:
            return {"success": True, "data": []}
        clause, clause_params = build_in_clause("company", companies)
        query += f" AND oc.Company_Id IN ({clause})"
        params.update(clause_params)

    if filters.get("Estatus"):
        query += " AND oc.Estatus = :estatus"
        params["estatus"] = str(filters["Estatus"])

    query += " ORDER BY oc.CreatedAt DESC"

    with get_connection() as connection:
        rows = connection.execute(text(query), params).mappings().all()
    return {"success": True, "data": [dict(row) for row in rows]}


def get_purchase_order(order_id: int, current_user: dict[str, Any]) -> dict[str, Any]:
    with get_connection() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden de compra no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        details = _load_purchase_details(connection, order_id)
        authorizations = _load_purchase_authorizations(connection, order_id)

    return {
        "success": True,
        "data": {
            **order,
            "detalle": details,
            "autorizaciones": authorizations,
        },
    }


def _update_product_cost(connection: Any, product_id: int | None, purchase_price: float | None) -> None:
    if not product_id or not purchase_price:
        return
    connection.execute(
        text(
            """
            UPDATE ERP_PRODUCTOS
            SET CostoInicial = :purchase_price, UpdatedAt = GETDATE()
            WHERE Producto_Id = :product_id
            """
        ),
        {"product_id": product_id, "purchase_price": purchase_price},
    )


def _replace_purchase_details(connection: Any, order_id: int, items: list[dict[str, Any]]) -> None:
    connection.execute(
        text("DELETE FROM ERP_COMPRA_ORDEN_DETALLE WHERE OC_Id = :oc_id"),
        {"oc_id": order_id},
    )
    for item in items:
        subtotal = float(item.get("Cantidad") or 0) * float(item.get("PrecioCompra") or 0)
        iva = subtotal * (float(item.get("IVA", DEFAULT_IVA_RATE) or DEFAULT_IVA_RATE) / 100)
        connection.execute(
            text(
                """
                INSERT INTO ERP_COMPRA_ORDEN_DETALLE (
                    OC_Id,
                    Producto_Id,
                    MateriaPrima_Id,
                    Descripcion,
                    Cantidad,
                    PrecioCompra,
                    Subtotal,
                    IVA,
                    Total
                )
                VALUES (
                    :oc_id,
                    :producto_id,
                    :materia_prima_id,
                    :descripcion,
                    :cantidad,
                    :precio_compra,
                    :subtotal,
                    :iva,
                    :total
                )
                """
            ),
            {
                "oc_id": order_id,
                "producto_id": int(item["Producto_Id"]) if item.get("Producto_Id") else None,
                "materia_prima_id": int(item["MateriaPrima_Id"]) if item.get("MateriaPrima_Id") else None,
                "descripcion": item.get("Descripcion") or "",
                "cantidad": float(item.get("Cantidad") or 0),
                "precio_compra": float(item.get("PrecioCompra") or 0),
                "subtotal": subtotal,
                "iva": iva,
                "total": subtotal + iva,
            },
        )
        _update_product_cost(
            connection,
            int(item["Producto_Id"]) if item.get("Producto_Id") else None,
            float(item.get("PrecioCompra") or 0),
        )


def create_purchase_order(payload: dict[str, Any], current_user: dict[str, Any]) -> dict[str, Any]:
    items = list(payload.get("items") or [])
    if not items:
        raise ApiServiceError(
            status_code=400,
            content={"success": False, "message": "La orden debe tener al menos un producto"},
        )

    company_id = int(payload.get("Company_Id") or 0)
    requisicion_id = int(payload.get("Requisicion_Id") or 0)
    _ensure_purchase_company_access(current_user, company_id)
    subtotal, total_iva, total = _calculate_purchase_totals(items)

    with get_transaction() as connection:
        requisicion = None
        observaciones = payload.get("Observaciones")
        if requisicion_id:
            requisicion = connection.execute(
                text(
                    """
                    SELECT Req_Id, Company_Id, NumeroReq, Estatus, OC_Id
                    FROM ERP_REQUISICION_COMPRA
                    WHERE Req_Id = :req_id
                    """
                ),
                {"req_id": requisicion_id},
            ).mappings().first()
            if not requisicion:
                raise ApiServiceError(
                    status_code=404,
                    content={"success": False, "message": "Requisicion no encontrada"},
                )
            _ensure_purchase_company_access(current_user, int(requisicion["Company_Id"]))
            if int(requisicion["Company_Id"]) != company_id:
                raise ApiServiceError(
                    status_code=400,
                    content={"success": False, "message": "La requisicion pertenece a otra empresa"},
                )
            if str(requisicion["Estatus"]) != "APROBADA":
                raise ApiServiceError(
                    status_code=400,
                    content={"success": False, "message": "La requisicion debe estar APROBADA para generar la OC"},
                )
            if requisicion.get("OC_Id"):
                raise ApiServiceError(
                    status_code=400,
                    content={"success": False, "message": "Esta requisicion ya fue convertida a una OC"},
                )

            origen = f"Generada desde Requisicion {requisicion['NumeroReq']}"
            observaciones_limpias = str(observaciones or "").strip()
            if not observaciones_limpias:
                observaciones = origen
            elif str(requisicion["NumeroReq"]) not in observaciones_limpias:
                observaciones = f"{observaciones_limpias}\n{origen}"

        numero_oc = _next_numero_oc(connection, company_id)
        inserted = connection.execute(
            text(
                """
                INSERT INTO ERP_COMPRA_ORDEN (
                    NumeroOC,
                    Company_Id,
                    Proveedor_Id,
                    FechaRequerida,
                    Moneda,
                    Subtotal,
                    IVA,
                    Total,
                    Estatus,
                    RequiereDobleAutorizacion,
                    FacturaReferencia,
                    Observaciones,
                    CreatedBy
                )
                OUTPUT INSERTED.OC_Id, INSERTED.NumeroOC
                VALUES (
                    :numero_oc,
                    :company_id,
                    :proveedor_id,
                    :fecha_requerida,
                    :moneda,
                    :subtotal,
                    :iva,
                    :total,
                    'BORRADOR',
                    :requiere_doble,
                    :factura_referencia,
                    :observaciones,
                    :created_by
                )
                """
            ),
            {
                "numero_oc": numero_oc,
                "company_id": company_id,
                "proveedor_id": int(payload.get("Proveedor_Id") or 0),
                "fecha_requerida": payload.get("FechaRequerida"),
                "moneda": payload.get("Moneda") or "MXN",
                "subtotal": subtotal,
                "iva": total_iva,
                "total": total,
                "requiere_doble": 0 if payload.get("RequiereDobleAutorizacion") is False else 1,
                "factura_referencia": payload.get("FacturaReferencia"),
                "observaciones": observaciones,
                "created_by": _username(current_user),
            },
        ).mappings().first()
        order_id = int(inserted["OC_Id"])
        _replace_purchase_details(connection, order_id, items)

        if requisicion:
            connection.execute(
                text(
                    """
                    UPDATE ERP_REQUISICION_COMPRA
                    SET Estatus = 'CONVERTIDA',
                        OC_Id = :oc_id,
                        UpdatedAt = GETDATE()
                    WHERE Req_Id = :req_id
                    """
                ),
                {"oc_id": order_id, "req_id": requisicion_id},
            )

    return {
        "success": True,
        "OC_Id": order_id,
        "NumeroOC": inserted["NumeroOC"],
        "Requisicion_Id": requisicion_id or None,
    }


def update_purchase_order(order_id: int, payload: dict[str, Any], current_user: dict[str, Any]) -> dict[str, Any]:
    items = list(payload.get("items") or [])
    with get_transaction() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        if str(order["Estatus"]) != "BORRADOR":
            raise ApiServiceError(
                status_code=400,
                content={"success": False, "message": "Solo se pueden editar ordenes en BORRADOR"},
            )

        subtotal, total_iva, total = _calculate_purchase_totals(items)
        connection.execute(
            text(
                """
                UPDATE ERP_COMPRA_ORDEN
                SET
                    Proveedor_Id = :proveedor_id,
                    FechaRequerida = :fecha_requerida,
                    Moneda = :moneda,
                    Subtotal = :subtotal,
                    IVA = :iva,
                    Total = :total,
                    RequiereDobleAutorizacion = :requiere_doble,
                    FacturaReferencia = :factura_referencia,
                    Observaciones = :observaciones,
                    UpdatedAt = GETDATE()
                WHERE OC_Id = :oc_id
                """
            ),
            {
                "oc_id": order_id,
                "proveedor_id": int(payload.get("Proveedor_Id") or 0),
                "fecha_requerida": payload.get("FechaRequerida"),
                "moneda": payload.get("Moneda") or "MXN",
                "subtotal": subtotal,
                "iva": total_iva,
                "total": total,
                "requiere_doble": 0 if payload.get("RequiereDobleAutorizacion") is False else 1,
                "factura_referencia": payload.get("FacturaReferencia"),
                "observaciones": payload.get("Observaciones"),
            },
        )
        _replace_purchase_details(connection, order_id, items)

    return {"success": True, "message": "Orden actualizada"}


def send_purchase_authorization(order_id: int, current_user: dict[str, Any]) -> dict[str, Any]:
    with get_transaction() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        if str(order["Estatus"]) != "BORRADOR":
            raise ApiServiceError(
                status_code=400,
                content={
                    "success": False,
                    "message": "Solo borradores pueden enviarse a autorizacion",
                },
            )
        connection.execute(
            text(
                """
                UPDATE ERP_COMPRA_ORDEN
                SET Estatus = 'PENDIENTE_AUTORIZACION', UpdatedAt = GETDATE()
                WHERE OC_Id = :oc_id
                """
            ),
            {"oc_id": order_id},
        )
    return {"success": True, "message": "Enviada a autorizacion"}


def authorize_purchase_order(
    order_id: int,
    payload: dict[str, Any],
    current_user: dict[str, Any],
) -> dict[str, Any]:
    nivel = int(payload.get("Nivel") or 0)
    aprobado = bool(payload.get("Aprobado"))
    comentarios = payload.get("Comentarios")
    if nivel not in {1, 2}:
        raise ApiServiceError(
            status_code=400,
            content={"success": False, "message": "Nivel debe ser 1 o 2"},
        )

    with get_transaction() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        if str(order["Estatus"]) != "PENDIENTE_AUTORIZACION":
            raise ApiServiceError(
                status_code=400,
                content={
                    "success": False,
                    "message": "La orden no esta pendiente de autorizacion",
                },
            )

        if nivel == 2:
            level_one = connection.execute(
                text(
                    """
                    SELECT Aprobado
                    FROM ERP_COMPRA_AUTORIZACION
                    WHERE OC_Id = :oc_id AND Nivel = 1
                    """
                ),
                {"oc_id": order_id},
            ).mappings().first()
            if not level_one or not bool(level_one["Aprobado"]):
                raise ApiServiceError(
                    status_code=400,
                    content={
                        "success": False,
                        "message": "El nivel 1 debe aprobar antes del nivel 2",
                    },
                )

        exists = connection.execute(
            text(
                """
                SELECT 1
                FROM ERP_COMPRA_AUTORIZACION
                WHERE OC_Id = :oc_id AND Nivel = :nivel
                """
            ),
            {"oc_id": order_id, "nivel": nivel},
        ).first()
        params = {
            "oc_id": order_id,
            "nivel": nivel,
            "user_id": int(current_user["User_Id"]) if current_user.get("User_Id") else None,
            "aprobado": 1 if aprobado else 0,
            "comentarios": comentarios,
        }
        if exists:
            connection.execute(
                text(
                    """
                    UPDATE ERP_COMPRA_AUTORIZACION
                    SET
                        User_Id = :user_id,
                        Aprobado = :aprobado,
                        FechaDecision = GETDATE(),
                        Comentarios = :comentarios
                    WHERE OC_Id = :oc_id AND Nivel = :nivel
                    """
                ),
                params,
            )
        else:
            connection.execute(
                text(
                    """
                    INSERT INTO ERP_COMPRA_AUTORIZACION (
                        OC_Id,
                        Nivel,
                        User_Id,
                        Aprobado,
                        Comentarios
                    )
                    VALUES (
                        :oc_id,
                        :nivel,
                        :user_id,
                        :aprobado,
                        :comentarios
                    )
                    """
                ),
                params,
            )

        if not aprobado:
            connection.execute(
                text(
                    """
                    UPDATE ERP_COMPRA_ORDEN
                    SET Estatus = 'RECHAZADA', UpdatedAt = GETDATE()
                    WHERE OC_Id = :oc_id
                    """
                ),
                {"oc_id": order_id},
            )
            return {"success": True, "message": "Orden rechazada"}

        auth_count = connection.execute(
            text(
                """
                SELECT COUNT(*) AS Total
                FROM ERP_COMPRA_AUTORIZACION
                WHERE OC_Id = :oc_id AND Aprobado = 1
                """
            ),
            {"oc_id": order_id},
        ).mappings().first()
        required = 2 if bool(order.get("RequiereDobleAutorizacion")) else 1
        total_auth = int((auth_count or {}).get("Total") or 0)
        if total_auth >= required:
            connection.execute(
                text(
                    """
                    UPDATE ERP_COMPRA_ORDEN
                    SET Estatus = 'AUTORIZADA', UpdatedAt = GETDATE()
                    WHERE OC_Id = :oc_id
                    """
                ),
                {"oc_id": order_id},
            )
            return {
                "success": True,
                "message": "Orden autorizada completamente",
                "autorizada": True,
            }

    return {
        "success": True,
        "message": f"Nivel {nivel} aprobado, pendiente siguiente nivel",
        "autorizada": False,
    }


def mark_purchase_order_bought(
    order_id: int,
    payload: dict[str, Any],
    current_user: dict[str, Any],
) -> dict[str, Any]:
    with get_transaction() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        if str(order["Estatus"]) != "AUTORIZADA":
            raise ApiServiceError(
                status_code=400,
                content={
                    "success": False,
                    "message": "La orden debe estar AUTORIZADA para registrar la compra",
                },
            )
        connection.execute(
            text(
                """
                UPDATE ERP_COMPRA_ORDEN
                SET
                    Estatus = 'COMPRADA',
                    FacturaReferencia = COALESCE(:factura_referencia, FacturaReferencia),
                    UpdatedAt = GETDATE()
                WHERE OC_Id = :oc_id
                """
            ),
            {"oc_id": order_id, "factura_referencia": payload.get("FacturaReferencia")},
        )

    return {"success": True, "message": "Compra registrada exitosamente"}


def upload_purchase_invoice(
    order_id: int,
    filename: str,
    content_type: str | None,
    file_bytes: bytes,
    current_user: dict[str, Any],
) -> dict[str, Any]:
    with get_transaction() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        if str(order["Estatus"]) not in {"AUTORIZADA", "COMPRADA"}:
            raise ApiServiceError(
                status_code=400,
                content={
                    "success": False,
                    "message": "Solo puede cargar factura en ordenes AUTORIZADAS o COMPRADAS",
                },
            )

        extension = Path(filename or "").suffix.lower() or ALLOWED_INVOICE_CONTENT_TYPES.get(content_type or "", ".pdf")
        if content_type not in ALLOWED_INVOICE_CONTENT_TYPES and extension not in {".pdf", ".xml", ".png", ".jpg", ".jpeg"}:
            raise ApiServiceError(
                status_code=400,
                content={
                    "success": False,
                    "message": "Formato invalido. Use PDF, XML, PNG o JPG",
                },
            )

        UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
        safe_numero = re.sub(r"[^a-zA-Z0-9_-]", "_", str(order.get("NumeroOC") or f"OC_{order_id}"))
        unique_name = f"factura_{safe_numero}_{order_id}_{len(file_bytes)}{extension}"
        file_path = UPLOADS_DIR / unique_name
        file_path.write_bytes(file_bytes)

        public_url = f"/uploads/compras-facturas/{unique_name}"
        connection.execute(
            text(
                """
                UPDATE ERP_COMPRA_ORDEN
                SET FacturaArchivoUrl = :factura_archivo_url, UpdatedAt = GETDATE()
                WHERE OC_Id = :oc_id
                """
            ),
            {"oc_id": order_id, "factura_archivo_url": public_url},
        )

    return {
        "success": True,
        "message": "Factura cargada exitosamente",
        "data": {"FacturaArchivoUrl": public_url},
    }


def receive_purchase_goods(
    order_id: int,
    payload: dict[str, Any],
    current_user: dict[str, Any],
) -> dict[str, Any]:
    warehouse_id = int(payload.get("Almacen_Id") or 0)
    items = list(payload.get("items") or [])
    if not warehouse_id:
        raise ApiServiceError(
            status_code=400,
            content={"success": False, "message": "Se requiere Almacen_Id"},
        )
    if not items:
        raise ApiServiceError(
            status_code=400,
            content={"success": False, "message": "Se requiere al menos una linea"},
        )

    with get_transaction() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        if str(order["Estatus"]) not in {"AUTORIZADA", "COMPRADA"}:
            raise ApiServiceError(
                status_code=400,
                content={
                    "success": False,
                    "message": (
                        "La OC debe estar en AUTORIZADA o COMPRADA para recibir mercancia "
                        f"(estado actual: {order['Estatus']})"
                    ),
                },
            )

        order_details = _load_purchase_details(connection, order_id)
        detail_map = {int(detail["OC_Detalle_Id"]): detail for detail in order_details}
        received_by = (
            f"{current_user.get('Name') or ''} {current_user.get('Lastname') or ''}".strip()
            or _username(current_user)
        )
        receipt = connection.execute(
            text(
                """
                INSERT INTO ERP_COMPRA_RECEPCION (
                    OC_Id,
                    Almacen_Id,
                    Observaciones,
                    RecibidoPor
                )
                OUTPUT INSERTED.Recepcion_Id
                VALUES (
                    :oc_id,
                    :almacen_id,
                    :observaciones,
                    :recibido_por
                )
                """
            ),
            {
                "oc_id": order_id,
                "almacen_id": warehouse_id,
                "observaciones": payload.get("Observaciones"),
                "recibido_por": received_by,
            },
        ).mappings().first()
        receipt_id = int(receipt["Recepcion_Id"])

        warehouse = connection.execute(
            text("SELECT Company_Id FROM ERP_ALMACENES WHERE Almacen_Id = :almacen_id"),
            {"almacen_id": warehouse_id},
        ).mappings().first()

        for item in items:
            detail_id = int(item.get("OC_Detalle_Id") or 0)
            cantidad_recibida = float(item.get("CantidadRecibida") or 0)
            if not detail_id or cantidad_recibida <= 0:
                continue
            detail = detail_map.get(detail_id)
            if not detail:
                continue

            connection.execute(
                text(
                    """
                    INSERT INTO ERP_COMPRA_RECEPCION_DETALLE (
                        Recepcion_Id,
                        OC_Detalle_Id,
                        Producto_Id,
                        MateriaPrima_Id,
                        Descripcion,
                        CantidadOrdenada,
                        CantidadRecibida,
                        PrecioCompra
                    )
                    VALUES (
                        :recepcion_id,
                        :oc_detalle_id,
                        :producto_id,
                        :materia_prima_id,
                        :descripcion,
                        :cantidad_ordenada,
                        :cantidad_recibida,
                        :precio_compra
                    )
                    """
                ),
                {
                    "recepcion_id": receipt_id,
                    "oc_detalle_id": detail_id,
                    "producto_id": detail.get("Producto_Id"),
                    "materia_prima_id": detail.get("MateriaPrima_Id"),
                    "descripcion": detail.get("Descripcion"),
                    "cantidad_ordenada": float(detail.get("Cantidad") or 0),
                    "cantidad_recibida": cantidad_recibida,
                    "precio_compra": float(detail.get("PrecioCompra") or 0),
                },
            )

            if detail.get("MateriaPrima_Id"):
                exists = connection.execute(
                    text(
                        """
                        SELECT Cantidad
                        FROM ERP_STOCK_MP
                        WHERE MateriaPrima_Id = :materia_prima_id AND Almacen_Id = :almacen_id
                        """
                    ),
                    {
                        "materia_prima_id": int(detail["MateriaPrima_Id"]),
                        "almacen_id": warehouse_id,
                    },
                ).mappings().first()
                if exists:
                    connection.execute(
                        text(
                            """
                            UPDATE ERP_STOCK_MP
                            SET Cantidad = Cantidad + :cantidad
                            WHERE MateriaPrima_Id = :materia_prima_id AND Almacen_Id = :almacen_id
                            """
                        ),
                        {
                            "materia_prima_id": int(detail["MateriaPrima_Id"]),
                            "almacen_id": warehouse_id,
                            "cantidad": cantidad_recibida,
                        },
                    )
                else:
                    connection.execute(
                        text(
                            """
                            INSERT INTO ERP_STOCK_MP (MateriaPrima_Id, Almacen_Id, Cantidad)
                            VALUES (:materia_prima_id, :almacen_id, :cantidad)
                            """
                        ),
                        {
                            "materia_prima_id": int(detail["MateriaPrima_Id"]),
                            "almacen_id": warehouse_id,
                            "cantidad": cantidad_recibida,
                        },
                    )
                connection.execute(
                    text(
                        """
                        UPDATE ERP_MATERIA_PRIMA
                        SET CostoUnitario = :costo_unitario,
                            FechaUltimoCosto = GETDATE()
                        WHERE MateriaPrima_Id = :materia_prima_id
                        """
                    ),
                    {
                        "materia_prima_id": int(detail["MateriaPrima_Id"]),
                        "costo_unitario": float(detail.get("PrecioCompra") or 0),
                    },
                )
            elif detail.get("Producto_Id"):
                product_exists = connection.execute(
                    text(
                        """
                        SELECT Cantidad
                        FROM ERP_STOCK
                        WHERE Producto_Id = :producto_id AND Almacen_Id = :almacen_id
                        """
                    ),
                    {"producto_id": int(detail["Producto_Id"]), "almacen_id": warehouse_id},
                ).mappings().first()
                if product_exists:
                    connection.execute(
                        text(
                            """
                            UPDATE ERP_STOCK
                            SET Cantidad = Cantidad + :cantidad
                            WHERE Producto_Id = :producto_id AND Almacen_Id = :almacen_id
                            """
                        ),
                        {
                            "producto_id": int(detail["Producto_Id"]),
                            "almacen_id": warehouse_id,
                            "cantidad": cantidad_recibida,
                        },
                    )
                else:
                    connection.execute(
                        text(
                            """
                            INSERT INTO ERP_STOCK (Producto_Id, Almacen_Id, Cantidad, Stock_Minimo)
                            VALUES (:producto_id, :almacen_id, :cantidad, 0)
                            """
                        ),
                        {
                            "producto_id": int(detail["Producto_Id"]),
                            "almacen_id": warehouse_id,
                            "cantidad": cantidad_recibida,
                        },
                    )
                _update_product_cost(
                    connection,
                    int(detail["Producto_Id"]),
                    float(detail.get("PrecioCompra") or 0),
                )

                if warehouse and warehouse.get("Company_Id"):
                    # Mantiene el consolidado coherente al sumar producto terminado recibido por compra.
                    connection.execute(
                        text(
                            """
                            MERGE ERP_INVENTARIO_ESTADO_PRODUCTO AS target
                            USING (
                                SELECT
                                    :company_id AS Company_Id,
                                    :producto_id AS Producto_Id,
                                    :almacen_id AS Almacen_Id
                            ) AS source
                            ON target.Company_Id = source.Company_Id
                               AND target.Producto_Id = source.Producto_Id
                               AND target.Almacen_Id = source.Almacen_Id
                            WHEN MATCHED THEN
                                UPDATE SET
                                    CantidadAlmacen = (
                                        SELECT ISNULL(SUM(Cantidad), 0)
                                        FROM ERP_STOCK
                                        WHERE Producto_Id = :producto_id AND Almacen_Id = :almacen_id
                                    ),
                                    FechaCorte = GETDATE()
                            WHEN NOT MATCHED THEN
                                INSERT (
                                    Company_Id,
                                    Producto_Id,
                                    Almacen_Id,
                                    CantidadAlmacen,
                                    CantidadEnMaquina,
                                    CantidadEntregadaProduccion,
                                    CantidadEnProceso,
                                    FechaCorte
                                )
                                VALUES (
                                    :company_id,
                                    :producto_id,
                                    :almacen_id,
                                    (
                                        SELECT ISNULL(SUM(Cantidad), 0)
                                        FROM ERP_STOCK
                                        WHERE Producto_Id = :producto_id AND Almacen_Id = :almacen_id
                                    ),
                                    0,
                                    0,
                                    0,
                                    GETDATE()
                                );
                            """
                        ),
                        {
                            "company_id": int(warehouse["Company_Id"]),
                            "producto_id": int(detail["Producto_Id"]),
                            "almacen_id": warehouse_id,
                        },
                    )

        if str(order["Estatus"]) == "AUTORIZADA":
            connection.execute(
                text(
                    """
                    UPDATE ERP_COMPRA_ORDEN
                    SET Estatus = 'COMPRADA', UpdatedAt = GETDATE()
                    WHERE OC_Id = :oc_id
                    """
                ),
                {"oc_id": order_id},
            )

        # Asiento contable automático al recibir mercancía
        total_recibido = sum(
            float(it.get("CantidadRecibida") or 0) * float((detail_map.get(int(it.get("OC_Detalle_Id") or 0)) or {}).get("PrecioCompra") or 0)
            for it in items
            if int(it.get("OC_Detalle_Id") or 0) in detail_map
        )
        if total_recibido > 0:
            try:
                ledger_service.post_recepcion_compra(
                    connection,
                    recepcion_id=receipt_id,
                    company_id=int(order["Company_Id"]),
                    total=total_recibido,
                )
            except Exception:
                pass  # best-effort

    return {
        "success": True,
        "message": "Mercancia recibida y stock actualizado",
        "data": {"Recepcion_Id": receipt_id},
    }


def list_purchase_receipts(order_id: int, current_user: dict[str, Any]) -> dict[str, Any]:
    with get_connection() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        receipts_rows = connection.execute(
            text(
                """
                SELECT r.*, a.Nombre AS Almacen
                FROM ERP_COMPRA_RECEPCION r
                LEFT JOIN ERP_ALMACENES a ON r.Almacen_Id = a.Almacen_Id
                WHERE r.OC_Id = :oc_id
                ORDER BY r.FechaRecepcion DESC
                """
            ),
            {"oc_id": order_id},
        ).mappings().all()

        receipts: list[dict[str, Any]] = []
        for receipt in receipts_rows:
            detail_rows = connection.execute(
                text(
                    """
                    SELECT
                        rd.*,
                        COALESCE(p.SKU, '') AS ProductoSKU,
                        COALESCE(p.Nombre, '') AS ProductoNombre,
                        COALESCE(mp.Nombre, '') AS MateriaPrimaNombre
                    FROM ERP_COMPRA_RECEPCION_DETALLE rd
                    LEFT JOIN ERP_PRODUCTOS p ON rd.Producto_Id = p.Producto_Id
                    LEFT JOIN ERP_MATERIA_PRIMA mp ON rd.MateriaPrima_Id = mp.MateriaPrima_Id
                    WHERE rd.Recepcion_Id = :recepcion_id
                    """
                ),
                {"recepcion_id": int(receipt["Recepcion_Id"])},
            ).mappings().all()
            receipts.append({**dict(receipt), "detalle": [dict(row) for row in detail_rows]})

    return {"success": True, "data": receipts}


def cancel_purchase_order(order_id: int, current_user: dict[str, Any]) -> dict[str, Any]:
    with get_transaction() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        if str(order["Estatus"]) in {"COMPRADA", "CANCELADA"}:
            raise ApiServiceError(
                status_code=400,
                content={
                    "success": False,
                    "message": f"No se puede cancelar una orden {order['Estatus']}",
                },
            )
        connection.execute(
            text(
                """
                UPDATE ERP_COMPRA_ORDEN
                SET Estatus = 'CANCELADA', UpdatedAt = GETDATE()
                WHERE OC_Id = :oc_id
                """
            ),
            {"oc_id": order_id},
        )
    return {"success": True, "message": "Orden cancelada"}


def get_purchase_order_pdf(order_id: int, current_user: dict[str, Any]) -> dict[str, Any]:
    with get_connection() as connection:
        order = _load_purchase_order(connection, order_id)
        if not order:
            raise ApiServiceError(
                status_code=404,
                content={"success": False, "message": "Orden no encontrada"},
            )
        _ensure_purchase_company_access(current_user, int(order["Company_Id"]))
        if str(order["Estatus"]) not in {"AUTORIZADA", "COMPRADA"}:
            raise ApiServiceError(
                status_code=403,
                content={
                    "success": False,
                    "message": "Solo se puede descargar el PDF cuando la OC esta autorizada o comprada",
                },
            )
        order["detalle"] = _load_purchase_details(connection, order_id)

    pdf_bytes = purchase_pdf_service.generar_pdf_orden_compra(order)
    numero_oc = str(order.get("NumeroOC") or f"oc-{order_id}")
    return {"filename": f"OC-{numero_oc}.pdf", "content": pdf_bytes}


def direct_purchase_registration(
    payload: dict[str, Any],
    current_user: dict[str, Any],
) -> dict[str, Any]:
    items = list(payload.get("items") or [])
    if not items:
        raise ApiServiceError(
            status_code=400,
            content={"success": False, "message": "Debe incluir al menos un producto"},
        )
    if not payload.get("FacturaReferencia"):
        raise ApiServiceError(
            status_code=400,
            content={
                "success": False,
                "message": "El registro directo requiere numero de factura",
            },
        )

    company_id = int(payload.get("Company_Id") or 0)
    _ensure_purchase_company_access(current_user, company_id)
    subtotal, total_iva, total = _calculate_purchase_totals(items)

    with get_transaction() as connection:
        numero_oc = _next_numero_oc(connection, company_id)
        inserted = connection.execute(
            text(
                """
                INSERT INTO ERP_COMPRA_ORDEN (
                    NumeroOC,
                    Company_Id,
                    Proveedor_Id,
                    Moneda,
                    Subtotal,
                    IVA,
                    Total,
                    Estatus,
                    RequiereDobleAutorizacion,
                    FacturaReferencia,
                    Observaciones,
                    CreatedBy
                )
                OUTPUT INSERTED.OC_Id, INSERTED.NumeroOC
                VALUES (
                    :numero_oc,
                    :company_id,
                    :proveedor_id,
                    :moneda,
                    :subtotal,
                    :iva,
                    :total,
                    'COMPRADA',
                    0,
                    :factura_referencia,
                    :observaciones,
                    :created_by
                )
                """
            ),
            {
                "numero_oc": numero_oc,
                "company_id": company_id,
                "proveedor_id": int(payload.get("Proveedor_Id") or 0),
                "moneda": payload.get("Moneda") or "MXN",
                "subtotal": subtotal,
                "iva": total_iva,
                "total": total,
                "factura_referencia": payload.get("FacturaReferencia"),
                "observaciones": payload.get("Observaciones"),
                "created_by": _username(current_user),
            },
        ).mappings().first()
        order_id = int(inserted["OC_Id"])
        _replace_purchase_details(connection, order_id, items)

    return {
        "success": True,
        "OC_Id": order_id,
        "NumeroOC": inserted["NumeroOC"],
        "message": "Compra registrada directamente",
    }
